"""Combined FedEx API and USPS website status updater."""
from __future__ import annotations

import json
import os
import re
import threading
import time
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlencode

import openpyxl
from openpyxl.styles import PatternFill

from fedex_credentials import has_credentials
from fedex_tracking import MAX_TRACKING_NUMBERS_PER_REQUEST, track_numbers

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except ImportError:
    sync_playwright = None
    PlaywrightTimeoutError = RuntimeError


FEDEX_SHEET_TO_PROFILE = {"UNI": "UNI", "FENIX": "FENIX"}
USPS_SHEETS = ("UNI", "EMBY", "FENIX", "SOL")

COL_DATE = 1
COL_CARRIER = 4
COL_TRACKING = 5
COL_REMARK = 6

DELIVERED_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")

FEDEX_RE = re.compile(
    r"\b(?:FEDEX|FED\s*EX|BX\s*FX|FX\s*S/O|FX\s*P/O|"
    r"STANDARD\s+OVERNIGHT|PRIORITY\s+OVERNIGHT|FIRST\s+OVERNIGHT)\b",
    re.I,
)
USPS_RE = re.compile(
    r"\b(?:USPS|U\.?S\.?\s*POSTAL|UNITED\s+STATES\s+POSTAL)\b",
    re.I,
)

USPS_BATCH_SIZE = 35
USPS_URL = "https://tools.usps.com/go/TrackConfirmAction"
USPS_TIMEOUT = 90000


def _settings_path():
    base = os.environ.get("LOCALAPPDATA")
    folder = (
        Path(base) / "UniCreation" / "ShipmentBot"
        if base else Path.home() / ".uni_creation_shipment_bot"
    )
    folder.mkdir(parents=True, exist_ok=True)
    return folder / "fedex_tracking_settings.json"


SETTINGS = _settings_path()


def load_tracking_time(default="16:00"):
    try:
        value = json.loads(
            SETTINGS.read_text(encoding="utf-8")
        ).get("fedex_tracking_time", default)
        return str(value).strip() or default
    except Exception:
        return default


def save_tracking_time(value):
    try:
        SETTINGS.write_text(
            json.dumps(
                {"fedex_tracking_time": str(value).strip()},
                indent=2,
            ),
            encoding="utf-8",
        )
    except Exception:
        pass


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in (
        "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d",
        "%d%b%y", "%d%b%Y",
    ):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except Exception:
            pass
    return None


def _current_month(value, today):
    return bool(
        value
        and value.year == today.year
        and value.month == today.month
    )


def _tracking(value):
    return "".join(
        ch for ch in str(value or "").upper()
        if ch.isalnum()
    )


def _batches(values, size):
    for index in range(0, len(values), size):
        yield values[index:index + size]


def _sheet(workbook, expected):
    actual = next(
        (
            name for name in workbook.sheetnames
            if name.strip().upper() == expected.upper()
        ),
        None,
    )
    return workbook[actual] if actual else None


def _normalize_status(text):
    value = re.sub(r"\s+", " ", str(text or "")).strip().upper()
    if not value:
        return "UNKNOWN"
    if "DELIVERED" in value:
        return "DELIVERED"
    if "OUT FOR DELIVERY" in value:
        return "OUT FOR DELIVERY"
    if any(x in value for x in (
        "DELIVERY ATTEMPT", "NOTICE LEFT", "NO ACCESS",
        "RECEPTACLE BLOCKED",
    )):
        return "DELIVERY ATTEMPTED"
    if any(x in value for x in (
        "ALERT", "EXCEPTION", "RETURN TO SENDER",
        "HELD AT POST OFFICE", "UNCLAIMED",
        "INSUFFICIENT ADDRESS",
    )):
        return "EXCEPTION"
    if any(x in value for x in (
        "IN TRANSIT", "MOVING THROUGH NETWORK",
        "ARRIVED AT USPS", "DEPARTED USPS",
        "PROCESSED THROUGH", "ARRIVING LATE",
    )):
        return "IN TRANSIT"
    if any(x in value for x in (
        "ACCEPTED", "USPS IN POSSESSION",
        "PICKED UP", "ORIGIN ACCEPTANCE",
    )):
        return "ACCEPTED"
    if any(x in value for x in (
        "LABEL CREATED", "PRE-SHIPMENT",
        "USPS AWAITING ITEM", "SHIPPING PARTNER",
    )):
        return "LABEL CREATED"
    if any(x in value for x in (
        "STATUS NOT AVAILABLE", "NOT TRACKABLE",
        "TRACKING NUMBER MAY BE INCORRECT",
    )):
        return "NOT FOUND"
    return value[:80]


def _set_status(ws, row, status):
    cell = ws.cell(row, COL_REMARK)
    cell.value = status
    if status == "DELIVERED":
        cell.fill = DELIVERED_FILL


def _open_workbook(path):
    for attempt in range(5):
        try:
            return openpyxl.load_workbook(path)
        except (PermissionError, OSError):
            if attempt == 4:
                raise PermissionError(
                    "Close the workbook in Excel and try again."
                )
            time.sleep(2)


def _update_fedex(workbook, environment, today, log):
    changed = False
    updated = unchanged = failed = 0
    log("[FEDEX] Starting FedEx API update.")

    for sheet_name, profile in FEDEX_SHEET_TO_PROFILE.items():
        ws = _sheet(workbook, sheet_name)
        if ws is None:
            log(f"[FEDEX] {sheet_name}: sheet not found; skipped.")
            continue

        rows = []
        for row in range(1, ws.max_row + 1):
            carrier = str(ws.cell(row, COL_CARRIER).value or "")
            number = _tracking(ws.cell(row, COL_TRACKING).value)
            old = str(ws.cell(row, COL_REMARK).value or "").strip().upper()
            shipped = _as_date(ws.cell(row, COL_DATE).value)

            if not FEDEX_RE.search(carrier):
                continue
            if not number or not _current_month(shipped, today):
                continue
            if old == "DELIVERED":
                _set_status(ws, row, "DELIVERED")
                changed = True
                continue
            rows.append((row, number, old))

        log(f"[FEDEX] {sheet_name}: {len(rows)} eligible shipment(s).")
        if not rows:
            continue
        if not has_credentials(profile):
            failed += len(rows)
            log(f"[FEDEX] {profile}: credentials missing.")
            continue

        results = {}
        unique = list(dict.fromkeys(number for _, number, _ in rows))
        for batch_no, batch in enumerate(
            _batches(unique, MAX_TRACKING_NUMBERS_PER_REQUEST), 1
        ):
            try:
                log(
                    f"[FEDEX] {profile}: batch {batch_no}, "
                    f"{len(batch)} number(s)."
                )
                results.update(
                    track_numbers(
                        profile,
                        batch,
                        environment=environment,
                        log=log,
                    )
                )
            except Exception as exc:
                failed += len(batch)
                log(
                    f"[FEDEX] Batch {batch_no} failed: "
                    f"{type(exc).__name__}: {exc}"
                )

        for row, number, old in rows:
            result = results.get(number)
            if not result:
                failed += 1
                continue
            status = _normalize_status(result.get("status"))
            if status == old:
                unchanged += 1
                continue
            _set_status(ws, row, status)
            changed = True
            updated += 1
            log(f"[FEDEX] {sheet_name} row {row}: {number} -> {status}")

    return {
        "changed": changed,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
    }


def _usps_url(numbers):
    query = urlencode({
        "tRef": "fullpage",
        "tLc": str(len(numbers)),
        "text28777": "",
        "tLabels": ",".join(numbers),
        "tABt": "false",
    })
    return f"{USPS_URL}?{query}"


def _extract_usps_results(page, numbers):
    """
    Wait for USPS results without using Page.wait_for_function().

    Some packaged Playwright installations can contain mismatched wrapper
    and implementation versions. A normal Python polling loop avoids that
    compatibility problem while still waiting for the dynamically rendered
    USPS results.
    """
    deadline = time.time() + (USPS_TIMEOUT / 1000)
    body = ""

    while time.time() < deadline:
        try:
            body = page.locator("body").inner_text(timeout=5000)
        except Exception:
            body = ""

        upper_body = body.upper()

        if any(number in body for number in numbers):
            break

        if any(
            phrase in upper_body
            for phrase in (
                "DELIVERED",
                "OUT FOR DELIVERY",
                "IN TRANSIT",
                "MOVING THROUGH NETWORK",
                "PRE-SHIPMENT",
                "LABEL CREATED",
                "USPS IN POSSESSION",
                "STATUS NOT AVAILABLE",
                "NOT TRACKABLE",
                "VERIFY YOU ARE HUMAN",
                "CAPTCHA",
                "ACCESS DENIED",
            )
        ):
            break

        time.sleep(1)

    if not body:
        body = page.locator("body").inner_text(timeout=30000)

    upper = body.upper()

    if any(x in upper for x in (
        "VERIFY YOU ARE HUMAN", "CAPTCHA", "ACCESS DENIED",
    )):
        raise RuntimeError(
            "USPS displayed human verification or access denied."
        )

    results = {}
    for number in numbers:
        text = ""
        try:
            locator = page.get_by_text(number, exact=False).first
            if locator.count():
                text = locator.evaluate(
                    """el => {
                        let node = el;
                        for (let i = 0; i < 8 && node; i++, node = node.parentElement) {
                            const t = (node.innerText || "").trim();
                            if (t.length >= 20 && t.length <= 2500 &&
                                /Delivered|Out for Delivery|In Transit|Moving Through Network|Pre-Shipment|Label Created|Accepted|USPS in Possession|Alert|Notice Left|Status Not Available|Not Trackable/i.test(t)) {
                                return t;
                            }
                        }
                        return (el.parentElement && el.parentElement.innerText) || el.innerText || "";
                    }"""
                )
        except Exception:
            text = ""

        if not text:
            index = body.find(number)
            if index >= 0:
                text = body[max(0, index - 200):index + 1800]

        results[number] = _normalize_status(text)

    return results


def _track_usps_batch(numbers, log):
    if sync_playwright is None:
        raise RuntimeError(
            "The playwright package is not included in this build."
        )

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            channel="msedge",
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
            ],
        )
        try:
            context = browser.new_context(
                locale="en-US",
                viewport={"width": 1440, "height": 1000},
            )
            page = context.new_page()
            page.set_default_timeout(30000)

            log(
                f"[USPS] Opening USPS tracking for "
                f"{len(numbers)} number(s)."
            )
            page.goto(
                _usps_url(numbers),
                wait_until="domcontentloaded",
                timeout=USPS_TIMEOUT,
            )
            try:
                page.wait_for_load_state("networkidle", timeout=20000)
            except PlaywrightTimeoutError:
                pass

            log("[USPS] Waiting for USPS result cards...")
            results = _extract_usps_results(page, numbers)
            log(
                f"[USPS] USPS page returned "
                f"{sum(1 for value in results.values() if value != 'UNKNOWN')} "
                f"usable status(es)."
            )
            return results
        finally:
            browser.close()


def _update_usps(workbook, today, log):
    changed = False
    updated = unchanged = failed = 0
    references = []

    log("[USPS] Starting USPS website update.")
    log("[USPS] Processing current-month rows in batches of 35.")

    for sheet_name in USPS_SHEETS:
        ws = _sheet(workbook, sheet_name)
        if ws is None:
            log(f"[USPS] {sheet_name}: sheet not found; skipped.")
            continue

        count = 0
        for row in range(1, ws.max_row + 1):
            carrier = str(ws.cell(row, COL_CARRIER).value or "")
            number = _tracking(ws.cell(row, COL_TRACKING).value)
            old = str(ws.cell(row, COL_REMARK).value or "").strip().upper()
            shipped = _as_date(ws.cell(row, COL_DATE).value)

            if not USPS_RE.search(carrier):
                continue
            if not number or not _current_month(shipped, today):
                continue
            if old == "DELIVERED":
                _set_status(ws, row, "DELIVERED")
                changed = True
                continue

            references.append((ws, sheet_name, row, number, old))
            count += 1

        log(f"[USPS] {sheet_name}: {count} eligible shipment(s).")

    unique = list(dict.fromkeys(item[3] for item in references))
    results = {}

    for batch_no, batch in enumerate(
        _batches(unique, USPS_BATCH_SIZE), 1
    ):
        try:
            log(
                f"[USPS] Batch {batch_no}: "
                f"{len(batch)} tracking number(s)."
            )
            results.update(_track_usps_batch(batch, log))
        except Exception as exc:
            log(
                f"[USPS] Batch {batch_no} failed: "
                f"{type(exc).__name__}: {exc}"
            )

    for ws, sheet_name, row, number, old in references:
        status = results.get(number)
        if not status or status == "UNKNOWN":
            failed += 1
            log(
                f"[USPS] {sheet_name} row {row}: "
                f"no usable status for {number}."
            )
            continue
        if status == old:
            unchanged += 1
            continue

        _set_status(ws, row, status)
        changed = True
        updated += 1
        log(f"[USPS] {sheet_name} row {row}: {number} -> {status}")

    return {
        "changed": changed,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
    }


def update_fedex_statuses(
    excel_path,
    *,
    environment="production",
    today=None,
    log=print,
):
    """
    Existing app.py can keep calling this function.
    It now updates both FedEx and USPS.
    """
    today = today or date.today()

    if not excel_path or not os.path.isfile(excel_path):
        message = f"[TRACKING] Workbook not found: {excel_path}"
        log(message)
        return {
            "updated": 0,
            "unchanged": 0,
            "failed": 0,
            "message": message,
        }

    log("[TRACKING] Starting combined FedEx + USPS update.")
    log(f"[TRACKING] Current month: {today.strftime('%B %Y')}")
    log("[TRACKING] Writing statuses to Remark column F.")

    workbook = _open_workbook(excel_path)
    try:
        fedex = _update_fedex(
            workbook,
            environment,
            today,
            log,
        )
        usps = _update_usps(
            workbook,
            today,
            log,
        )

        if fedex["changed"] or usps["changed"]:
            workbook.save(excel_path)
            log("[TRACKING] Workbook saved successfully.")
        else:
            log("[TRACKING] No workbook changes were required.")

        updated = fedex["updated"] + usps["updated"]
        unchanged = fedex["unchanged"] + usps["unchanged"]
        failed = fedex["failed"] + usps["failed"]

        message = (
            f"[TRACKING] Finished: updated={updated}, "
            f"unchanged={unchanged}, failed={failed}. "
            f"FedEx updated={fedex['updated']}; "
            f"USPS updated={usps['updated']}."
        )
        log(message)

        return {
            "updated": updated,
            "unchanged": unchanged,
            "failed": failed,
            "fedex": fedex,
            "usps": usps,
            "message": message,
        }
    finally:
        workbook.close()


update_all_tracking_statuses = update_fedex_statuses


class FedExStatusScheduler(threading.Thread):
    """Existing name retained; scheduled runs now update FedEx and USPS."""

    def __init__(
        self,
        get_excel_path,
        get_target,
        log=print,
        *,
        environment="production",
        poll_seconds=20,
    ):
        super().__init__(daemon=True)
        self.get_excel_path = get_excel_path
        self.get_target = get_target
        self.log = log
        self.environment = environment
        self.poll_seconds = poll_seconds
        self._stop_event = threading.Event()
        self._last_run_date = None

    def stop(self):
        self._stop_event.set()

    def run(self):
        self.log("FedEx + USPS status scheduler started.")

        while not self._stop_event.is_set():
            now = datetime.now()
            target = str(self.get_target() or "").strip()

            if (
                target
                and now.strftime("%H:%M") == target
                and self._last_run_date != now.date()
            ):
                self._last_run_date = now.date()
                path = str(self.get_excel_path() or "").strip()

                if path and os.path.isfile(path):
                    self.log(
                        "[TRACKING] Running scheduled "
                        "FedEx + USPS update..."
                    )
                    update_fedex_statuses(
                        path,
                        environment=self.environment,
                        today=now.date(),
                        log=self.log,
                    )
                else:
                    self.log(
                        "[TRACKING] Scheduled update skipped: "
                        "Excel path is missing."
                    )

            self._stop_event.wait(self.poll_seconds)

        self.log("FedEx + USPS status scheduler stopped.")