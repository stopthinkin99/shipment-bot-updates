import datetime
import os
import time

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ------------------------------------------------------------------ #
#  CONFIG  (all the knobs you might want to tweak live up here)
# ------------------------------------------------------------------ #

# Fallback sheet if the parser didn't decide one (it normally does)
DEFAULT_SHEET = "FENIX"

# Column layout (1-indexed) --------------------------------------- #
COL_DATE     = 1   # A
COL_SHIP_TO  = 2   # B
COL_INVOICE  = 3   # C
COL_CARRIER  = 4   # D
COL_TRACKING = 5   # E
COL_REMARK   = 7   # G  (e.g. "Zales Account")

# How wide the blue divider row should be (A..G = 7) --------------- #
DIVIDER_COLUMNS = 7

# Your sheet sometimes has ONE blank white row before the blue
# divider (see rows 75->76 / 84->85 in your screenshot). Flip this
# to False if you want the blue divider directly under the last row.
BLANK_ROW_BEFORE_DIVIDER = True

# If a value appears more than this many rows below the previous
# real value, treat it as stray junk (this is what kills the
# "row 711" jump). Your date groups only have 1-2 blank rows, so 15
# is very safe.
MAX_GAP = 15

# FENIX sheet displays dates as 7/9/2026. We write a REAL Excel date
# (so it sorts/filters correctly) and format the cell as m/d/yyyy.
# Set WRITE_DATE_AS_TEXT True only if you want literal text instead.
WRITE_DATE_AS_TEXT = False
DATE_NUMBER_FORMAT = "m/d/yyyy"   # -> 7/9/2026
DATE_OUTPUT_FORMAT = "%d%b%y"     # only used if WRITE_DATE_AS_TEXT is True

# Save retry behaviour (file is often locked because it's open in Excel)
SAVE_RETRIES = 6      # how many times to retry a locked save
SAVE_DELAY   = 5      # seconds to wait between retries

BLUE_FILL = PatternFill(
    start_color="5B9BD5",
    end_color="5B9BD5",
    fill_type="solid",
)


# ------------------------------------------------------------------ #
#  DATE HELPERS
# ------------------------------------------------------------------ #

def _coerce_to_date(value):
    """
    Turn whatever is in the cell OR whatever OCR gave us into a
    plain datetime.date, so two dates can actually be compared.
    Returns None if it can't be parsed.
    """
    if value is None:
        return None

    # Real Excel date/datetime cell
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    text = str(value).strip()
    if text == "":
        return None

    # strptime is case-insensitive for month names, so "09JUL26" is fine
    formats = [
        "%d%b%y",     # 09JUL26
        "%d%b%Y",     # 09JUL2026
        "%m/%d/%Y",   # 7/9/2026
        "%m/%d/%y",   # 7/9/26
        "%Y-%m-%d",   # 2026-07-09
    ]

    for fmt in formats:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _format_date_for_cell(d):
    """Format a date the way the FENIX sheet stores it, e.g. 09JUL26."""
    if WRITE_DATE_AS_TEXT:
        return d.strftime(DATE_OUTPUT_FORMAT).upper()
    return d   # write a real date object


# ------------------------------------------------------------------ #
#  LAST-ROW DETECTION  (the actual bug fix)
# ------------------------------------------------------------------ #

def _find_last_data_row(ws):
    """
    Find the last row that holds REAL shipment data.

    A row counts only if its DATE cell (col A):
      - is not None, AND
      - is not a blank / whitespace-only string.

    We also bail out if we hit a value that sits more than MAX_GAP
    rows below the previous real value -- that's stray content far
    past the table (the reason data was landing at row 711).
    """
    last_real = 0

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, COL_DATE).value

        # skip truly empty and empty-string cells
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue

        # big jump below the table => stray junk, stop here
        if last_real != 0 and (row - last_real) > MAX_GAP:
            break

        last_real = row

    return last_real


# ------------------------------------------------------------------ #
#  SAVE / LOCK HANDLING
# ------------------------------------------------------------------ #

def _is_file_locked(file_path):
    """
    True if the workbook looks like it's open in Excel.
    Excel drops a hidden lock file called ~$<name> next to the real
    file, and also holds an exclusive OS lock. We check both.
    """
    lock_file = os.path.join(
        os.path.dirname(file_path),
        "~$" + os.path.basename(file_path),
    )
    if os.path.exists(lock_file):
        return True

    # Fallback: try to open for append. If Excel has it, this raises.
    try:
        with open(file_path, "a"):
            return False
    except PermissionError:
        return True


def _safe_save(wb, file_path):
    """
    Save with retries. If the file is locked (open in Excel) we wait
    and try again instead of crashing the watcher thread.
    Returns True on success, False if it stayed locked.
    """
    for attempt in range(1, SAVE_RETRIES + 1):
        try:
            wb.save(file_path)
            return True
        except PermissionError:
            print(
                f"[WARN] '{os.path.basename(file_path)}' is locked "
                f"(attempt {attempt}/{SAVE_RETRIES}). "
                f"Likely open in Excel. Retrying in {SAVE_DELAY}s..."
            )
            time.sleep(SAVE_DELAY)

    print("[ERROR] Save failed — file stayed locked. Shipment NOT written.")
    return False


# ------------------------------------------------------------------ #
#  MAIN
# ------------------------------------------------------------------ #

def update_excel(file_path, data):

    wb = load_workbook(file_path, keep_links=False)

    sheet_name = data.get("sheet") or DEFAULT_SHEET
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    last_row = _find_last_data_row(ws)

    last_date = (
        _coerce_to_date(ws.cell(last_row, COL_DATE).value)
        if last_row else None
    )
    new_date = _coerce_to_date(data["date"])

    if new_date is None:
        raise ValueError(f"Could not parse OCR date: {data['date']!r}")

    # ---- decide where to write ---- #
    if last_row == 0:
        # empty sheet safety net
        write_row = 1

    elif last_date == new_date:
        # same day -> stack directly under the last shipment
        write_row = last_row + 1

    else:
        # new day -> (optional blank row) + blue divider + data
        divider_row = last_row + 2 if BLANK_ROW_BEFORE_DIVIDER else last_row + 1

        for col in range(1, DIVIDER_COLUMNS + 1):
            ws.cell(divider_row, col).fill = BLUE_FILL

        write_row = divider_row + 1

    # ---- debug ---- #
    print("[DEBUG] Last data row  :", last_row)
    print("[DEBUG] Last date      :", last_date)
    print("[DEBUG] New date       :", new_date)
    print("[DEBUG] Writing at row :", write_row)

    # ---- write the shipment ---- #
    date_cell = ws.cell(write_row, COL_DATE)
    date_cell.value = _format_date_for_cell(new_date)
    if not WRITE_DATE_AS_TEXT:
        date_cell.number_format = DATE_NUMBER_FORMAT

    ws.cell(write_row, COL_SHIP_TO).value  = data.get("ship_to")
    ws.cell(write_row, COL_INVOICE).value  = data.get("invoice")
    ws.cell(write_row, COL_CARRIER).value  = data.get("carrier")
    ws.cell(write_row, COL_TRACKING).value = data.get("tracking_number")

    # Remark (column G) — e.g. "Zales Account". Only write if present.
    if data.get("remark"):
        ws.cell(write_row, COL_REMARK).value = data["remark"]

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 60), 12)

    if not _safe_save(wb, file_path):
        raise PermissionError(
            f"Could not save '{file_path}'. Close it in Excel and try again."
        )

    print(f"[SUCCESS] Shipment added at row {write_row}")
    return write_row
