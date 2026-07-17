"""Update the Excel REMARK column with current shipment status."""
from __future__ import annotations

import os
import re
import time
from collections import defaultdict
from typing import Any

from openpyxl import load_workbook

from tracking_service import TrackingServiceError, get_tracking_status, normalize_carrier, normalize_tracking_number

DEFAULT_SHEETS = ("UNI", "EMBY", "FENIX", "SOL")
HEADER_SCAN_LIMIT = 60
SAVE_RETRIES = 6
SAVE_DELAY = 5
TERMINAL_STATUSES = {"DELIVERED", "RETURN TO SENDER", "CANCELLED"}


def _header(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _find_columns(ws):
    for row in range(1, min(ws.max_row, HEADER_SCAN_LIMIT) + 1):
        found = {}
        for col in range(1, ws.max_column + 1):
            value = _header(ws.cell(row, col).value)
            if value == "CARRIER":
                found["carrier"] = col
            elif value in {"TRACKINGNUMBER", "TRACKINGNO", "TRACKING"}:
                found["tracking"] = col
            elif value in {"REMARK", "REMARKS", "STATUS", "DELIVERYSTATUS"}:
                found["remark"] = col
        if {"carrier", "tracking", "remark"}.issubset(found):
            return row, found
    return None


def _save(wb, path: str) -> None:
    for attempt in range(1, SAVE_RETRIES + 1):
        try:
            wb.save(path)
            return
        except PermissionError:
            if attempt == SAVE_RETRIES:
                raise
            time.sleep(SAVE_DELAY)


def update_tracking_statuses(
    excel_path: str,
    *,
    sheets: tuple[str, ...] = DEFAULT_SHEETS,
    api_key: str | None = None,
    log_fn=print,
    skip_delivered: bool = True,
) -> dict[str, int]:
    if not excel_path:
        raise ValueError("Excel path is blank.")
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Workbook not found: {excel_path}")

    wb = load_workbook(excel_path, keep_links=False)
    counters = {"checked": 0, "updated": 0, "unchanged": 0, "skipped": 0, "errors": 0}
    targets = defaultdict(list)

    try:
        titles = {name.upper(): name for name in wb.sheetnames}

        for requested in sheets:
            actual = titles.get(requested.upper())
            if not actual:
                log_fn(f"[TRACKING] Sheet missing: {requested}")
                continue

            ws = wb[actual]
            found = _find_columns(ws)
            if not found:
                log_fn(f"[TRACKING] CARRIER/TRACKING/REMARK headers not found on {actual}")
                counters["errors"] += 1
                continue

            header_row, cols = found
            for row in range(header_row + 1, ws.max_row + 1):
                carrier = normalize_carrier(ws.cell(row, cols["carrier"]).value)
                tracking = normalize_tracking_number(ws.cell(row, cols["tracking"]).value)
                current = str(ws.cell(row, cols["remark"]).value or "").strip().upper()

                if not tracking:
                    continue
                if not carrier:
                    counters["skipped"] += 1
                    continue
                if skip_delivered and current in TERMINAL_STATUSES:
                    counters["skipped"] += 1
                    continue

                targets[(carrier, tracking)].append((ws, row, cols["remark"]))

        for (carrier, tracking), rows in targets.items():
            counters["checked"] += 1
            try:
                result = get_tracking_status(carrier, tracking, api_key=api_key)
                status = result["status"]
                log_fn(f"[TRACKING] {carrier} {tracking}: {status}")
            except TrackingServiceError as exc:
                counters["errors"] += 1
                log_fn(f"[TRACKING ERROR] {carrier} {tracking}: {exc}")
                continue

            for ws, row, remark_col in rows:
                cell = ws.cell(row, remark_col)
                if str(cell.value or "").strip().upper() == status:
                    counters["unchanged"] += 1
                else:
                    cell.value = status
                    counters["updated"] += 1

        if counters["updated"]:
            _save(wb, excel_path)
            log_fn(f"[TRACKING] Saved workbook; {counters['updated']} status cell(s) updated.")
        else:
            log_fn("[TRACKING] No status changes to save.")

        return counters
    finally:
        wb.close()