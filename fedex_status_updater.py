"""Combined FedEx API and USPS website status updater."""
from __future__ import annotations

import json
import os
import re
import threading
import time
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlencode

import openpyxl
from openpyxl.styles import PatternFill

from fedex_credentials import has_credentials
from fedex_tracking import MAX_TRACKING_NUMBERS_PER_REQUEST, track_numbers

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except ImportError:
    sync_playwright = None
    PlaywrightTimeoutError = RuntimeError


FEDEX_SHEET_TO_PROFILE = {"UNI": "UNI", "FENIX": "FENIX"}
USPS_SHEETS = ("UNI", "EMBY", "FENIX", "SOL")

COL_DATE = 1
COL_CARRIER = 4
COL_TRACKING = 5
COL_REMARK = 6

DELIVERED_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")

FEDEX_RE = re.compile(
    r"(?:"
    r"\b(?:FEDEX|FED\s*EX|BX\s*FX|FX\s*S/O|FX\s*P/O|"
    r"STANDARD\s+OVERNIGHT|PRIORITY\s+OVERNIGHT|FIRST\s+OVERNIGHT)\b"
    r"|^\s*M\s*/\s*E\s*$"
    r")",
    re.I,
)
USPS_RE = re.compile(
    r"\b(?:USPS|U\.?S\.?\s*POSTAL|UNITED\s+STATES\s+POSTAL)\b",
    re.I,
)

MALCA_RE = re.compile(
    r"\b(?:MALCA\s*[- ]?\s*AMIT|MALKA\s*[- ]?\s*AMIT)\b",
    re.I,
)
MALCA_SHEETS = ("UNI", "EMBY", "FENIX", "SOL")
MALCA_TRACKING_URL = "https://tracking.malca-amit.com/"
MALCA_TIMEOUT = 60000

UPS_RE = re.compile(r"\bUPS\b", re.I)
UPS_SHEETS = ("UNI", "EMBY", "FENIX", "SOL")
UPS_TRACKING_URL = "https://www.ups.com/track"
UPS_TIMEOUT = 60000
UPS_HUMAN_VERIFICATION_TIMEOUT = 600000
UPS_BATCH_SIZE = 25

USPS_BATCH_SIZE = 35
USPS_URL = "https://tools.usps.com/go/TrackConfirmAction"
USPS_TIMEOUT = 90000
USPS_HUMAN_VERIFICATION_TIMEOUT = 600000  # 10 minutes


def _settings_path():
    base = os.environ.get("LOCALAPPDATA")
    folder = (
        Path(base) / "UniCreation" / "ShipmentBot"
        if base else Path.home() / ".uni_creation_shipment_bot"
    )
    folder.mkdir(parents=True, exist_ok=True)
    return folder / "fedex_tracking_settings.json"


SETTINGS = _settings_path()


def load_tracking_time(default="16:00"):
    try:
        value = json.loads(
            SETTINGS.read_text(encoding="utf-8")
        ).get("fedex_tracking_time", default)
        return str(value).strip() or default
    except Exception:
        return default


def save_tracking_time(value):
    try:
        SETTINGS.write_text(
            json.dumps(
                {"fedex_tracking_time": str(value).strip()},
                indent=2,
            ),
            encoding="utf-8",
        )
    except Exception:
        pass


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in (
        "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d",
        "%d%b%y", "%d%b%Y",
    ):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except Exception:
            pass
    return None


def _current_month(value, today):
    return bool(
        value
        and value.year == today.year
        and value.month == today.month
    )


def _tracking(value):
    return "".join(
        ch for ch in str(value or "").upper()
        if ch.isalnum()
    )


def _batches(values, size):
    for index in range(0, len(values), size):
        yield values[index:index + size]


def _sheet(workbook, expected):
    actual = next(
        (
            name for name in workbook.sheetnames
            if name.strip().upper() == expected.upper()
        ),
        None,
    )
    return workbook[actual] if actual else None


def _normalize_status(text):
    """
    Normalize FedEx API status text into stable Excel values.

    FedEx may return short labels, detailed localized descriptions, or
    event-code-derived wording. More specific phrases are checked before
    broader ones so, for example, "Ready for Pickup" is not treated as
    "Picked Up".
    """
    value = re.sub(r"\s+", " ", str(text or "")).strip().upper()

    if not value:
        return "UNKNOWN"

    # Final delivery
    if any(phrase in value for phrase in (
        "DELIVERED",
        "DELIVERY COMPLETED",
        "PROOF OF DELIVERY",
    )):
        return "DELIVERED"

    # Final-mile delivery
    if any(phrase in value for phrase in (
        "OUT FOR DELIVERY",
        "ON FEDEX VEHICLE FOR DELIVERY",
        "VEHICLE DISPATCHED",
        "ENROUTE TO DELIVERY",
    )):
        return "OUT FOR DELIVERY"

    # Pickup / hold options
    if any(phrase in value for phrase in (
        "READY FOR PICKUP",
        "READY FOR RECIPIENT PICKUP",
        "HOLD AT LOCATION REQUEST ACCEPTED",
        "AVAILABLE FOR PICKUP",
        "AT FEDEX LOCATION",
    )):
        return "AVAILABLE FOR PICKUP"

    # Delivery attempt
    if any(phrase in value for phrase in (
        "DELIVERY ATTEMPT",
        "DELIVERY EXCEPTION - CUSTOMER NOT AVAILABLE",
        "CUSTOMER NOT AVAILABLE",
        "RECIPIENT NOT AVAILABLE",
        "BUSINESS CLOSED",
        "NO ONE AVAILABLE",
    )):
        return "DELIVERY ATTEMPTED"

    # Customs / regulatory
    if any(phrase in value for phrase in (
        "CLEARANCE DELAY",
        "CUSTOMS DELAY",
        "CLEARANCE EXCEPTION",
        "REGULATORY AGENCY CLEARANCE DELAY",
    )):
        return "CUSTOMS DELAY"

    if any(phrase in value for phrase in (
        "CLEARANCE IN PROGRESS",
        "CUSTOMS CLEARANCE IN PROGRESS",
        "INTERNATIONAL SHIPMENT RELEASE",
        "CLEARED CUSTOMS",
        "EXPORT APPROVED",
    )):
        return "CUSTOMS PROCESSING"

    # Returns
    if any(phrase in value for phrase in (
        "RETURNING PACKAGE TO SHIPPER",
        "RETURN TO SHIPPER",
        "RETURN TO SENDER",
        "RETURN SHIPMENT",
        "RETURN INITIATED",
    )):
        return "RETURN TO SENDER"

    # Exceptions and delays
    if any(phrase in value for phrase in (
        "SHIPMENT EXCEPTION",
        "DELIVERY EXCEPTION",
        "EXCEPTION",
        "WEATHER DELAY",
        "OPERATIONAL DELAY",
        "DELIVERY DELAY",
        "DELAYED",
        "ADDRESS CORRECTED",
        "ADDRESS CHANGE REQUESTED",
        "LOCATION CHANGED",
    )):
        return "EXCEPTION"

    # In transit
    if any(phrase in value for phrase in (
        "ON THE WAY",
        "IN TRANSIT",
        "AT LOCAL FEDEX FACILITY",
        "AT FEDEX FACILITY",
        "AT SORT FACILITY",
        "ARRIVED AT FEDEX LOCATION",
        "DEPARTED FEDEX LOCATION",
        "LEFT FEDEX ORIGIN FACILITY",
        "PLANE IN FLIGHT",
        "PLANE LANDED",
        "TRANSFER",
        "IN PROGRESS",
        "ARRIVED AT PORT OF ENTRY",
    )):
        return "IN TRANSIT"

    # Possession / pickup
    if any(phrase in value for phrase in (
        "PICKED UP",
        "PICKED UP - SEE DETAILS",
        "IN FEDEX POSSESSION",
        "DROPPED OFF",
        "TENDERED AT FEDEX LOCATION",
        "FEDEX HAS THE PACKAGE",
    )):
        return "PICKED UP"

    # Shipment created but not yet physically received
    if any(phrase in value for phrase in (
        "SHIPMENT INFORMATION SENT TO FEDEX",
        "LABEL CREATED",
        "SHIPMENT CREATED",
        "ORDER CREATED",
        "FEDEX AWAITING PACKAGE",
        "PACKAGE DATA TRANSMITTED TO FEDEX",
    )):
        return "LABEL CREATED"

    # Cancelled / invalid
    if any(phrase in value for phrase in (
        "SHIPMENT CANCELLED",
        "SHIPMENT CANCELED",
        "CANCELLED",
        "CANCELED",
    )):
        return "CANCELLED"

    if any(phrase in value for phrase in (
        "TRACKING NUMBER NOT FOUND",
        "NO RECORD OF THIS TRACKING NUMBER",
        "INVALID TRACKING NUMBER",
        "NOT FOUND",
    )):
        return "NOT FOUND"

    # Preserve an unfamiliar current FedEx description rather than mapping
    # it incorrectly. Excel receives a concise value for review.
    return value[:80]

def _normalize_malca_status(text):
    """
    Normalize Malca-Amit result text.

    Malca-Amit does not publish a complete public status-code dictionary.
    The parser therefore recognizes broad secure-logistics wording and
    preserves unfamiliar concise statuses rather than guessing.
    """
    value = re.sub(r"\s+", " ", str(text or "")).strip().upper()

    if not value:
        return "UNKNOWN"

    if any(phrase in value for phrase in (
        "HAWB NUMBER NOT FOUND",
        "SHIPMENT NUMBER NOT FOUND",
        "REFERENCE NUMBER NOT FOUND",
        "NO SHIPMENT FOUND",
        "NO RECORD FOUND",
    )):
        return "NOT FOUND"

    if any(phrase in value for phrase in (
        "DELIVERED",
        "DELIVERY COMPLETED",
        "PROOF OF DELIVERY",
        "POD RECEIVED",
        "SIGNED FOR",
        "CONSIGNEE RECEIVED",
    )):
        return "DELIVERED"

    if any(phrase in value for phrase in (
        "OUT FOR DELIVERY",
        "WITH COURIER",
        "COURIER OUT FOR DELIVERY",
        "DISPATCHED FOR DELIVERY",
        "EN ROUTE FOR DELIVERY",
    )):
        return "OUT FOR DELIVERY"

    if any(phrase in value for phrase in (
        "DELIVERY ATTEMPT",
        "UNABLE TO DELIVER",
        "CONSIGNEE CLOSED",
        "CONSIGNEE UNAVAILABLE",
        "NO ANSWER",
        "DELIVERY FAILED",
    )):
        return "DELIVERY ATTEMPTED"

    if any(phrase in value for phrase in (
        "AVAILABLE FOR COLLECTION",
        "AVAILABLE FOR PICKUP",
        "READY FOR COLLECTION",
        "READY FOR PICKUP",
        "HELD FOR COLLECTION",
    )):
        return "AVAILABLE FOR PICKUP"

    if any(phrase in value for phrase in (
        "CUSTOMS HOLD",
        "CUSTOMS DELAY",
        "AWAITING CUSTOMS CLEARANCE",
        "CLEARANCE IN PROGRESS",
        "UNDER CUSTOMS CLEARANCE",
    )):
        return "CUSTOMS PROCESSING"

    if any(phrase in value for phrase in (
        "RETURN TO SHIPPER",
        "RETURN TO SENDER",
        "RETURNED TO ORIGIN",
        "RETURNED",
    )):
        return "RETURN TO SENDER"

    if any(phrase in value for phrase in (
        "EXCEPTION",
        "ON HOLD",
        "SECURITY HOLD",
        "DELAY",
        "DELAYED",
        "ADDRESS ISSUE",
        "DOCUMENTATION REQUIRED",
        "AWAITING INSTRUCTIONS",
    )):
        return "EXCEPTION"

    if any(phrase in value for phrase in (
        "IN TRANSIT",
        "DEPARTED",
        "ARRIVED",
        "RECEIVED AT",
        "FORWARDED",
        "TRANSFERRED",
        "AT HUB",
        "AT BRANCH",
        "EN ROUTE",
        "IN PROCESS",
    )):
        return "IN TRANSIT"

    if any(phrase in value for phrase in (
        "PICKED UP",
        "COLLECTED",
        "SHIPMENT RECEIVED",
        "RECEIVED FROM SHIPPER",
        "BOOKED",
        "ACCEPTED",
    )):
        return "PICKED UP"

    if any(phrase in value for phrase in (
        "SHIPMENT CREATED",
        "BOOKING CREATED",
        "AWAITING COLLECTION",
        "AWAITING PICKUP",
        "INFORMATION RECEIVED",
    )):
        return "LABEL CREATED"

    if any(phrase in value for phrase in (
        "CANCELLED",
        "CANCELED",
        "SHIPMENT VOIDED",
    )):
        return "CANCELLED"

    return value[:80]

def _normalize_ups_status(text):
    """
    Normalize UPS tracking-page wording.

    UPS officially groups statuses into Label Created, Shipped/On the Way,
    Out for Delivery, Transferred to Post Office, Delivered, Delivered to
    a UPS Access Point, and Exception. Common detailed page wording is also
    covered below.
    """
    value = re.sub(r"\s+", " ", str(text or "")).strip().upper()

    if not value:
        return "UNKNOWN"

    if any(phrase in value for phrase in (
        "TRACKING INFORMATION IS NOT YET AVAILABLE",
        "WE COULD NOT LOCATE THE SHIPMENT DETAILS",
        "TRACKING NUMBER IS NOT VALID",
        "INVALID TRACKING NUMBER",
        "DETAILS NOT YET AVAILABLE",
    )):
        return "NOT FOUND"

    # Access Point is not final delivery to the recipient.
    if any(phrase in value for phrase in (
        "DELIVERED TO A UPS ACCESS POINT",
        "DELIVERED TO UPS ACCESS POINT",
        "READY FOR CUSTOMER PICKUP",
        "READY FOR PICKUP",
        "AVAILABLE FOR PICKUP",
        "HELD FOR PICKUP",
    )):
        return "AVAILABLE FOR PICKUP"

    if any(phrase in value for phrase in (
        "DELIVERED",
        "DELIVERY COMPLETE",
        "PROOF OF DELIVERY",
        "LEFT AT",
        "RECEIVED BY",
    )):
        return "DELIVERED"

    if any(phrase in value for phrase in (
        "OUT FOR DELIVERY",
        "LOADED ON DELIVERY VEHICLE",
        "DESTINATION SCAN",
    )):
        return "OUT FOR DELIVERY"

    if any(phrase in value for phrase in (
        "TRANSFERRED TO POST OFFICE FOR DELIVERY",
        "TRANSFERRED TO THE POST OFFICE",
        "PACKAGE TRANSFERRED TO POST OFFICE",
        "POST OFFICE ATTEMPTED DELIVERY",
    )):
        return "TRANSFERRED TO USPS"

    if any(phrase in value for phrase in (
        "DELIVERY ATTEMPT",
        "WE MISSED YOU",
        "RECEIVER WAS NOT AVAILABLE",
        "RECEIVER UNAVAILABLE",
        "BUSINESS WAS CLOSED",
        "CUSTOMER WAS NOT AVAILABLE",
        "UPS DELIVERY NOTICE",
    )):
        return "DELIVERY ATTEMPTED"

    if any(phrase in value for phrase in (
        "RETURNING TO SENDER",
        "RETURN TO SENDER",
        "RETURN SERVICE",
        "RETURNED TO SENDER",
    )):
        return "RETURN TO SENDER"

    if any(phrase in value for phrase in (
        "EXCEPTION",
        "DELAY",
        "WEATHER MAY CAUSE A DELAY",
        "SEVERE WEATHER",
        "ADDRESS INFORMATION REQUIRED",
        "ADDRESS NEEDS CORRECTION",
        "HELD",
        "DAMAGE REPORTED",
        "PACKAGE DAMAGED",
        "CUSTOMS DELAY",
        "CLEARANCE INFORMATION REQUIRED",
        "MECHANICAL FAILURE",
        "EMERGENCY SITUATION",
    )):
        return "EXCEPTION"

    if any(phrase in value for phrase in (
        "SHIPPED/ON THE WAY",
        "ON THE WAY",
        "IN TRANSIT",
        "DEPARTED FROM FACILITY",
        "ARRIVED AT FACILITY",
        "PROCESSING AT UPS FACILITY",
        "ORIGIN SCAN",
        "DEPARTURE SCAN",
        "ARRIVAL SCAN",
        "IMPORT SCAN",
        "EXPORT SCAN",
        "WAREHOUSE SCAN",
        "YOUR PACKAGE IS ON THE WAY",
    )):
        return "IN TRANSIT"

    if any(phrase in value for phrase in (
        "WE HAVE YOUR PACKAGE",
        "PICKED UP",
        "PICKUP SCAN",
        "DROP-OFF",
        "UPS HAS POSSESSION",
    )):
        return "PICKED UP"

    if any(phrase in value for phrase in (
        "LABEL CREATED",
        "SHIPPER CREATED A LABEL",
        "UPS DOESN'T HAVE POSSESSION",
        "UPS DOES NOT HAVE POSSESSION",
        "SHIPMENT READY FOR UPS",
        "BILLING INFORMATION RECEIVED",
        "ORDER PROCESSED: READY FOR UPS",
    )):
        return "LABEL CREATED"

    if any(phrase in value for phrase in (
        "CANCELLED",
        "CANCELED",
        "VOIDED",
    )):
        return "CANCELLED"

    return value[:80]

def _set_status(ws, row, status):
    cell = ws.cell(row, COL_REMARK)
    cell.value = status
    if status == "DELIVERED":
        cell.fill = DELIVERED_FILL


def _open_workbook(path):
    for attempt in range(5):
        try:
            return openpyxl.load_workbook(path)
        except (PermissionError, OSError):
            if attempt == 4:
                raise PermissionError(
                    "Close the workbook in Excel and try again."
                )
            time.sleep(2)


def _update_fedex(workbook, environment, today, log):
    changed = False
    updated = unchanged = failed = 0
    log("[FEDEX] Starting FedEx API update.")

    for sheet_name, profile in FEDEX_SHEET_TO_PROFILE.items():
        ws = _sheet(workbook, sheet_name)
        if ws is None:
            log(f"[FEDEX] {sheet_name}: sheet not found; skipped.")
            continue

        rows = []
        for row in range(1, ws.max_row + 1):
            carrier = str(ws.cell(row, COL_CARRIER).value or "")
            number = _tracking(ws.cell(row, COL_TRACKING).value)
            old = str(ws.cell(row, COL_REMARK).value or "").strip().upper()
            shipped = _as_date(ws.cell(row, COL_DATE).value)

            if not FEDEX_RE.search(carrier):
                continue
            if not number or not _current_month(shipped, today):
                continue
            if old == "DELIVERED":
                _set_status(ws, row, "DELIVERED")
                changed = True
                continue
            rows.append((row, number, old))

        log(f"[FEDEX] {sheet_name}: {len(rows)} eligible shipment(s).")
        if not rows:
            continue
        if not has_credentials(profile):
            failed += len(rows)
            log(f"[FEDEX] {profile}: credentials missing.")
            continue

        results = {}
        unique = list(dict.fromkeys(number for _, number, _ in rows))
        for batch_no, batch in enumerate(
            _batches(unique, MAX_TRACKING_NUMBERS_PER_REQUEST), 1
        ):
            try:
                log(
                    f"[FEDEX] {profile}: batch {batch_no}, "
                    f"{len(batch)} number(s)."
                )
                results.update(
                    track_numbers(
                        profile,
                        batch,
                        environment=environment,
                        log=log,
                    )
                )
            except Exception as exc:
                failed += len(batch)
                log(
                    f"[FEDEX] Batch {batch_no} failed: "
                    f"{type(exc).__name__}: {exc}"
                )

        for row, number, old in rows:
            result = results.get(number)
            if not result:
                failed += 1
                continue
            status = _normalize_status(result.get("status"))
            log(
                f"[FEDEX] Parsed {number}: {status}"
            )
            if status == old:
                unchanged += 1
                continue
            _set_status(ws, row, status)
            changed = True
            updated += 1
            log(f"[FEDEX] {sheet_name} row {row}: {number} -> {status}")

    return {
        "changed": changed,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
    }


def _usps_url(numbers):
    query = urlencode({
        "tRef": "fullpage",
        "tLc": str(len(numbers)),
        "text28777": "",
        "tLabels": ",".join(numbers),
        "tABt": "false",
    })
    return f"{USPS_URL}?{query}"


def _normalize_usps_card_status(card_text, tracking_number):
    """
    Extract the CURRENT USPS status from one bounded shipment card.

    USPS pages contain a progress bar with generic stages:
        Pre-Shipment / In Transit / Out for Delivery / Delivered

    Those stage labels are not necessarily the current status. Therefore:
    - active/non-final statuses take priority;
    - DELIVERED requires a detailed delivery statement;
    - a bare progress-bar word "Delivered" is ignored unless it appears
      immediately as the current heading and no active status is present.
    """
    raw = re.sub(r"\r", "", str(card_text or ""))
    upper = raw.upper()

    number_index = upper.find(tracking_number.upper())
    if number_index >= 0:
        raw = raw[number_index + len(tracking_number):]
        upper = upper[number_index + len(tracking_number):]

    raw = raw[:2200]
    upper = upper[:2200]

    lines = [
        re.sub(r"\s+", " ", line).strip()
        for line in raw.splitlines()
        if line.strip()
    ]
    progress_only = {
        "PRE-SHIPMENT",
        "IN TRANSIT",
        "OUT FOR DELIVERY",
        "DELIVERED",
    }
    useful_lines = [
        line for line in lines
        if line.upper().strip(" :-") not in progress_only
    ]
    useful = "\n".join(useful_lines)
    useful_upper = useful.upper()

    active_patterns = [
        (r"\bOUT FOR REDELIVERY\b", "OUT FOR DELIVERY"),
        (r"\bOUT FOR DELIVERY\b", "OUT FOR DELIVERY"),
        (r"\bON THE WAY\b", "IN TRANSIT"),
        (r"\bMOVING THROUGH NETWORK\b", "IN TRANSIT"),
        (r"\bIN[- ]TRANSIT,\s*ARRIVING LATE\b", "IN TRANSIT - DELAYED"),
        (r"\bIN[- ]TRANSIT,\s*ARRIVING ON TIME\b", "IN TRANSIT"),
        (r"\bIN TRANSIT TO NEXT FACILITY\b", "IN TRANSIT"),
        (r"\bIN TRANSIT\b", "IN TRANSIT"),
        (r"\bARRIVED AT USPS REGIONAL FACILITY\b", "IN TRANSIT"),
        (r"\bDEPARTED USPS REGIONAL FACILITY\b", "IN TRANSIT"),
        (r"\bARRIVED AT USPS FACILITY\b", "IN TRANSIT"),
        (r"\bDEPARTED USPS FACILITY\b", "IN TRANSIT"),
        (r"\bPROCESSED THROUGH USPS FACILITY\b", "IN TRANSIT"),
        (r"\bARRIVED AT HUB\b", "IN TRANSIT"),
        (r"\bDEPARTED POST OFFICE\b", "IN TRANSIT"),
        (r"\bARRIVED AT POST OFFICE\b", "PREPARING FOR DELIVERY"),
        (r"\bPREPARING FOR DELIVERY\b", "PREPARING FOR DELIVERY"),
        (r"\bAVAILABLE FOR REDELIVERY OR PICKUP\b", "AVAILABLE FOR PICKUP"),
        (r"\bAVAILABLE FOR PICKUP\b", "AVAILABLE FOR PICKUP"),
        (r"\bHELD AT POST OFFICE\b", "AVAILABLE FOR PICKUP"),
        (r"\bREADY FOR PICKUP\b", "AVAILABLE FOR PICKUP"),
        (r"\bREDELIVERY SCHEDULED\b", "REDELIVERY SCHEDULED"),
        (r"\bPREPARED FOR REDELIVERY\b", "PREPARING FOR REDELIVERY"),
        (r"\bNOTICE LEFT\b", "DELIVERY ATTEMPTED"),
        (r"\bDELIVERY ATTEMPT(?:ED)?\b", "DELIVERY ATTEMPTED"),
        (r"\bNO ACCESS TO DELIVERY LOCATION\b", "DELIVERY ATTEMPTED"),
        (r"\bRECEPTACLE BLOCKED\b", "DELIVERY ATTEMPTED"),
        (r"\bANIMAL INTERFERENCE\b", "DELIVERY ATTEMPTED"),
        (r"\bBUSINESS CLOSED\b", "DELIVERY ATTEMPTED"),
        (r"\bAWAITING DELIVERY SCAN\b", "AWAITING DELIVERY SCAN"),
        (r"\bDELIVERY STATUS NOT UPDATED\b", "AWAITING DELIVERY SCAN"),
        (r"\bFORWARDED\b", "FORWARDED"),
        (r"\bUSPS IN POSSESSION OF ITEM\b", "ACCEPTED"),
        (r"\bSHIPMENT RECEIVED,\s*PACKAGE ACCEPTANCE PENDING\b", "ACCEPTANCE PENDING"),
        (r"\bPACKAGE ACCEPTANCE PENDING\b", "ACCEPTANCE PENDING"),
        (r"\bACCEPTED AT USPS ORIGIN FACILITY\b", "ACCEPTED"),
        (r"\bORIGIN ACCEPTANCE\b", "ACCEPTED"),
        (r"\bARRIVED SHIPPING PARTNER FACILITY,\s*USPS AWAITING ITEM\b", "SHIPPING PARTNER"),
        (r"\bDEPARTED SHIPPING PARTNER FACILITY,\s*USPS AWAITING ITEM\b", "SHIPPING PARTNER"),
        (r"\bPICKED UP BY SHIPPING PARTNER,\s*USPS AWAITING ITEM\b", "SHIPPING PARTNER"),
        (r"\bON ITS WAY TO USPS\b", "SHIPPING PARTNER"),
        (r"\bSHIPPING PARTNER\b", "SHIPPING PARTNER"),
        (r"\bPRE-SHIPMENT INFO SENT TO USPS,\s*USPS AWAITING ITEM\b", "LABEL CREATED"),
        (r"\bSHIPPING LABEL CREATED,\s*USPS AWAITING ITEM\b", "LABEL CREATED"),
        (r"\bUSPS AWAITING ITEM\b", "LABEL CREATED"),
        (r"\bPRE-SHIPMENT\b", "LABEL CREATED"),
        (r"\bINSUFFICIENT ADDRESS\b", "RETURN TO SENDER"),
        (r"\bNO SUCH NUMBER\b", "RETURN TO SENDER"),
        (r"\bADDRESSEE UNKNOWN\b", "RETURN TO SENDER"),
        (r"\bVACANT\b", "RETURN TO SENDER"),
        (r"\bUNCLAIMED\b", "RETURN TO SENDER"),
        (r"\bREFUSED\b", "RETURN TO SENDER"),
        (r"\bFORWARD EXPIRED\b", "RETURN TO SENDER"),
        (r"\bRETURN TO SENDER\b", "RETURN TO SENDER"),
        (r"\bMAIL RECOVERY CENTER\b", "EXCEPTION"),
        (r"\bALERT\b", "EXCEPTION"),
        (r"\bSTATUS NOT AVAILABLE\b", "NOT FOUND"),
        (r"\bNOT TRACKABLE\b", "NOT FOUND"),
        (r"\bTRACKING NUMBER MAY BE INCORRECT\b", "NOT FOUND"),
    ]

    active_matches = []
    for pattern, normalized in active_patterns:
        match = re.search(pattern, useful_upper, re.I)
        if match:
            active_matches.append(
                (match.start(), -len(match.group(0)), normalized)
            )

    if active_matches:
        active_matches.sort(key=lambda item: (item[0], item[1]))
        return active_matches[0][2]

    delivered_patterns = [
        r"\bYOUR ITEM WAS DELIVERED\b",
        r"\bYOUR PACKAGE WAS DELIVERED\b",
        r"\bDELIVERED,\s*(?:FRONT DESK|RECEPTION|MAIL ROOM|MAILROOM|"
        r"PARCEL LOCKER|PO BOX|GARAGE|PORCH|DOOR|INDIVIDUAL|AGENT|"
        r"LEFT WITH|RECEIVED BY|IN/AT MAILBOX)\b",
        r"\bDELIVERED TO AGENT(?: FOR FINAL DELIVERY)?\b",
        r"\bDELIVERED TO POSTAL AGENT\b",
        r"\bDELIVERED AT\b",
        r"\bDELIVERED ON\b",
        r"\bPICKED UP AT POST OFFICE\b",
        r"\bPICKED UP BY INDIVIDUAL AT POST OFFICE\b",
    ]

    for pattern in delivered_patterns:
        if re.search(pattern, useful_upper, re.I):
            return "DELIVERED"

    bare = re.search(r"\bDELIVERED\b", useful_upper[:160], re.I)
    if bare:
        return "DELIVERED"

    return "UNKNOWN"

def _usps_card_segments(body, numbers):
    """
    Split the complete USPS page text into one bounded segment per tracking
    number. Each segment ends where the next requested tracking number begins.
    """
    body_upper = body.upper()
    occurrences = []

    for number in numbers:
        start = body_upper.find(number.upper())
        if start >= 0:
            occurrences.append((start, number))

    occurrences.sort()
    segments = {}

    for index, (start, number) in enumerate(occurrences):
        end = (
            occurrences[index + 1][0]
            if index + 1 < len(occurrences)
            else min(len(body), start + 2500)
        )
        segments[number] = body[start:end]

    return segments


def _extract_usps_results(page, numbers, log):
    """
    Assisted USPS mode.

    The browser remains visible. If USPS displays a human-verification
    challenge, the employee completes it in Edge. The program then
    automatically continues when tracking results appear.
    """
    normal_deadline = time.time() + (USPS_TIMEOUT / 1000)
    verification_deadline = None
    verification_announced = False
    body = ""

    result_phrases = (
        "DELIVERED",
        "OUT FOR DELIVERY",
        "IN TRANSIT",
        "MOVING THROUGH NETWORK",
        "PRE-SHIPMENT",
        "LABEL CREATED",
        "USPS IN POSSESSION",
        "ACCEPTED",
        "STATUS NOT AVAILABLE",
        "NOT TRACKABLE",
    )
    challenge_phrases = (
        "VERIFY YOU ARE HUMAN",
        "CAPTCHA",
        "HUMAN VERIFICATION",
        "SECURITY CHECK",
        "PRESS AND HOLD",
        "ACCESS DENIED",
    )

    while True:
        try:
            body = page.locator("body").inner_text(timeout=5000)
        except Exception:
            body = ""

        upper_body = body.upper()
        challenge_present = any(
            phrase in upper_body for phrase in challenge_phrases
        )
        results_present = (
            any(number in body for number in numbers)
            and any(phrase in upper_body for phrase in result_phrases)
        )

        if results_present:
            if verification_announced:
                log(
                    "[USPS] Human verification completed. "
                    "Tracking results detected; continuing automatically."
                )
            break

        if challenge_present:
            if not verification_announced:
                verification_announced = True
                verification_deadline = (
                    time.time()
                    + (USPS_HUMAN_VERIFICATION_TIMEOUT / 1000)
                )
                log(
                    "[USPS] Human verification is required in the "
                    "open Microsoft Edge window."
                )
                log(
                    "[USPS] Complete the verification manually. "
                    "Do not close Edge; the program will continue "
                    "automatically afterward."
                )

            if time.time() >= verification_deadline:
                raise RuntimeError(
                    "USPS human verification was not completed "
                    "within 10 minutes."
                )
        else:
            if not verification_announced and time.time() >= normal_deadline:
                raise RuntimeError(
                    "USPS tracking results did not load within 90 seconds."
                )

        time.sleep(1)

    # Parse each shipment from its own bounded section of the page.
    # Do not use broad parent containers because they may contain several
    # shipment cards and progress labels such as "Delivered".
    segments = _usps_card_segments(body, numbers)
    results = {}

    for number in numbers:
        card_text = segments.get(number, "")
        status = _normalize_usps_card_status(
            card_text,
            number,
        )
        results[number] = status

        log(
            f"[USPS] Parsed {number}: {status}"
        )

    return results


def _track_usps_batch(numbers, log):
    if sync_playwright is None:
        raise RuntimeError(
            "The playwright package is not included in this build."
        )

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            channel="msedge",
            headless=False,
            args=[
                "--disable-dev-shm-usage",
                "--start-maximized",
            ],
        )
        try:
            context = browser.new_context(
                locale="en-US",
                viewport=None,
            )
            page = context.new_page()
            page.set_default_timeout(30000)

            log(
                f"[USPS] Opening visible USPS tracking window for "
                f"{len(numbers)} number(s)."
            )
            log(
                "[USPS] Tracking numbers are being submitted "
                "automatically."
            )
            page.goto(
                _usps_url(numbers),
                wait_until="domcontentloaded",
                timeout=USPS_TIMEOUT,
            )
            try:
                page.wait_for_load_state("networkidle", timeout=20000)
            except PlaywrightTimeoutError:
                pass

            log("[USPS] Waiting for USPS results or human verification...")
            results = _extract_usps_results(page, numbers, log)
            log(
                f"[USPS] USPS page returned "
                f"{sum(1 for value in results.values() if value != 'UNKNOWN')} "
                f"usable status(es)."
            )
            return results
        finally:
            browser.close()


def _update_usps(workbook, today, log):
    changed = False
    updated = unchanged = failed = 0
    references = []

    log("[USPS] Starting USPS website update.")
    log("[USPS] Processing current-month rows in batches of 35.")

    for sheet_name in USPS_SHEETS:
        ws = _sheet(workbook, sheet_name)
        if ws is None:
            log(f"[USPS] {sheet_name}: sheet not found; skipped.")
            continue

        count = 0
        for row in range(1, ws.max_row + 1):
            carrier = str(ws.cell(row, COL_CARRIER).value or "")
            number = _tracking(ws.cell(row, COL_TRACKING).value)
            old = str(ws.cell(row, COL_REMARK).value or "").strip().upper()
            shipped = _as_date(ws.cell(row, COL_DATE).value)

            if not USPS_RE.search(carrier):
                continue
            if not number or not _current_month(shipped, today):
                continue
            if old == "DELIVERED":
                _set_status(ws, row, "DELIVERED")
                changed = True
                continue

            references.append((ws, sheet_name, row, number, old))
            count += 1

        log(f"[USPS] {sheet_name}: {count} eligible shipment(s).")

    unique = list(dict.fromkeys(item[3] for item in references))
    results = {}

    for batch_no, batch in enumerate(
        _batches(unique, USPS_BATCH_SIZE), 1
    ):
        try:
            log(
                f"[USPS] Batch {batch_no}: "
                f"{len(batch)} tracking number(s)."
            )
            results.update(_track_usps_batch(batch, log))
        except Exception as exc:
            log(
                f"[USPS] Batch {batch_no} failed: "
                f"{type(exc).__name__}: {exc}"
            )

    for ws, sheet_name, row, number, old in references:
        status = results.get(number)
        if not status or status == "UNKNOWN":
            failed += 1
            log(
                f"[USPS] {sheet_name} row {row}: "
                f"no usable status for {number}."
            )
            continue
        if status == old:
            unchanged += 1
            continue

        _set_status(ws, row, status)
        changed = True
        updated += 1
        log(f"[USPS] {sheet_name} row {row}: {number} -> {status}")

    return {
        "changed": changed,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
    }



def _malca_find_tracking_input(page):
    """Find the first visible shipment/reference input on the tracking page."""
    candidates = [
        page.get_by_label(
            re.compile(r"shipment|reference|hawb", re.I)
        ),
        page.locator(
            'input[placeholder*="shipment" i], '
            'input[placeholder*="reference" i], '
            'input[name*="hawb" i], '
            'input[id*="hawb" i], '
            'input[type="text"]'
        ),
    ]

    for candidate in candidates:
        try:
            count = candidate.count()
            for index in range(count):
                item = candidate.nth(index)
                if item.is_visible():
                    return item
        except Exception:
            continue

    return None


def _malca_click_find(page):
    candidates = [
        page.get_by_role(
            "button",
            name=re.compile(r"^find$|track|search", re.I),
        ),
        page.locator(
            'button:has-text("Find"), '
            'input[type="submit"], '
            'button[type="submit"]'
        ),
    ]

    for candidate in candidates:
        try:
            count = candidate.count()
            for index in range(count):
                item = candidate.nth(index)
                if item.is_visible():
                    item.click()
                    return True
        except Exception:
            continue

    return False


def _malca_page_requires_more_information(body):
    upper = str(body or "").upper()
    return (
        "TO CONTINUE, PLEASE PROVIDE NEXT INFORMATION" in upper
        or (
            "ORIGIN COUNTRY" in upper
            and "ACTIVITY TYPE" in upper
        )
    )


def _extract_malca_status_from_body(body, tracking_number):
    """
    Inspect the rendered result page. Avoid treating page headings and form
    labels as shipment statuses.
    """
    cleaned = re.sub(r"\s+", " ", str(body or "")).strip()
    upper = cleaned.upper()

    if any(phrase in upper for phrase in (
        "HAWB NUMBER NOT FOUND",
        "SHIPMENT NUMBER NOT FOUND",
        "REFERENCE NUMBER NOT FOUND",
        "NO SHIPMENT FOUND",
    )):
        return "NOT FOUND"

    # Prefer strong status phrases found anywhere in the rendered result.
    return _normalize_malca_status(cleaned)


def _track_malca_number(page, tracking_number, log):
    page.goto(
        MALCA_TRACKING_URL,
        wait_until="domcontentloaded",
        timeout=MALCA_TIMEOUT,
    )

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass

    tracking_input = _malca_find_tracking_input(page)
    if tracking_input is None:
        raise RuntimeError(
            "Could not find the Malca-Amit tracking-number field."
        )

    tracking_input.fill(tracking_number)

    if not _malca_click_find(page):
        tracking_input.press("Enter")

    deadline = time.time() + (MALCA_TIMEOUT / 1000)
    body = ""

    while time.time() < deadline:
        try:
            body = page.locator("body").inner_text(timeout=5000)
        except Exception:
            body = ""

        upper = body.upper()

        if _malca_page_requires_more_information(body):
            return {
                "status": None,
                "reason": "additional_information_required",
            }

        if any(phrase in upper for phrase in (
            "DELIVERED",
            "OUT FOR DELIVERY",
            "IN TRANSIT",
            "PICKED UP",
            "COLLECTED",
            "EXCEPTION",
            "ON HOLD",
            "DELAY",
            "HAWB NUMBER NOT FOUND",
            "SHIPMENT NUMBER NOT FOUND",
            "REFERENCE NUMBER NOT FOUND",
            "NO SHIPMENT FOUND",
        )):
            status = _extract_malca_status_from_body(
                body,
                tracking_number,
            )
            return {
                "status": status,
                "reason": None,
            }

        time.sleep(1)

    return {
        "status": None,
        "reason": "timeout",
    }


def _update_malca(workbook, today, log):
    changed = False
    updated = unchanged = failed = needs_info = 0
    references = []

    log("[MALCA-AMIT] Starting automated website update.")
    log(
        "[MALCA-AMIT] Processing current-month shipments "
        "one tracking number at a time."
    )

    for sheet_name in MALCA_SHEETS:
        ws = _sheet(workbook, sheet_name)
        if ws is None:
            log(
                f"[MALCA-AMIT] {sheet_name}: sheet not found; skipped."
            )
            continue

        count = 0
        for row in range(1, ws.max_row + 1):
            carrier = str(ws.cell(row, COL_CARRIER).value or "")
            number = _tracking(ws.cell(row, COL_TRACKING).value)
            old = str(
                ws.cell(row, COL_REMARK).value or ""
            ).strip().upper()
            shipped = _as_date(ws.cell(row, COL_DATE).value)

            if not MALCA_RE.search(carrier):
                continue
            if not number or not _current_month(shipped, today):
                continue
            if old == "DELIVERED":
                _set_status(ws, row, "DELIVERED")
                changed = True
                continue

            references.append((ws, sheet_name, row, number, old))
            count += 1

        log(
            f"[MALCA-AMIT] {sheet_name}: "
            f"{count} eligible shipment(s)."
        )

    if not references:
        return {
            "changed": changed,
            "updated": updated,
            "unchanged": unchanged,
            "failed": failed,
            "needs_info": needs_info,
        }

    if sync_playwright is None:
        log(
            "[MALCA-AMIT] Playwright is not available in this build."
        )
        return {
            "changed": changed,
            "updated": updated,
            "unchanged": unchanged,
            "failed": len(references),
            "needs_info": needs_info,
        }

    # Track each unique number once and reuse the result for duplicate rows.
    unique_numbers = list(
        dict.fromkeys(item[3] for item in references)
    )
    results = {}

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            channel="msedge",
            headless=True,
            args=["--disable-dev-shm-usage"],
        )
        try:
            context = browser.new_context(
                locale="en-US",
                viewport={"width": 1440, "height": 1000},
            )
            page = context.new_page()
            page.set_default_timeout(30000)

            for index, number in enumerate(unique_numbers, start=1):
                try:
                    log(
                        f"[MALCA-AMIT] Tracking {index}/"
                        f"{len(unique_numbers)}: {number}"
                    )
                    result = _track_malca_number(
                        page,
                        number,
                        log,
                    )
                    results[number] = result

                    if result.get("status"):
                        log(
                            f"[MALCA-AMIT] {number} -> "
                            f"{result['status']}"
                        )
                    elif (
                        result.get("reason")
                        == "additional_information_required"
                    ):
                        log(
                            f"[MALCA-AMIT] {number}: website requires "
                            "origin country/activity type; row left unchanged."
                        )
                    else:
                        log(
                            f"[MALCA-AMIT] {number}: no usable status "
                            "was returned."
                        )
                except Exception as exc:
                    results[number] = {
                        "status": None,
                        "reason": "error",
                    }
                    log(
                        f"[MALCA-AMIT] {number} failed: "
                        f"{type(exc).__name__}: {exc}"
                    )
        finally:
            browser.close()

    for ws, sheet_name, row, number, old in references:
        result = results.get(number) or {}
        status = result.get("status")
        reason = result.get("reason")

        if reason == "additional_information_required":
            needs_info += 1
            continue

        if not status or status == "UNKNOWN":
            failed += 1
            log(
                f"[MALCA-AMIT] {sheet_name} row {row}: "
                f"no usable status for {number}."
            )
            continue

        if status == old:
            unchanged += 1
            continue

        _set_status(ws, row, status)
        changed = True
        updated += 1
        log(
            f"[MALCA-AMIT] {sheet_name} row {row}: "
            f"{number} -> {status}"
        )

    return {
        "changed": changed,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
        "needs_info": needs_info,
    }



def _ups_tracking_url(numbers):
    return (
        f"{UPS_TRACKING_URL}?"
        + urlencode({
            "loc": "en_US",
            "tracknum": ",".join(numbers),
            "requester": "ST",
        })
    )


def _ups_challenge_present(body):
    upper = str(body or "").upper()
    return any(x in upper for x in (
        "VERIFY YOU ARE HUMAN",
        "CAPTCHA",
        "SECURITY CHECK",
        "ACCESS DENIED",
        "PRESS AND HOLD",
        "LET US KNOW YOU ARE HUMAN",
    ))


def _ups_result_phrases_present(body):
    upper = str(body or "").upper()
    return any(x in upper for x in (
        "DELIVERED",
        "OUT FOR DELIVERY",
        "ON THE WAY",
        "IN TRANSIT",
        "LABEL CREATED",
        "WE HAVE YOUR PACKAGE",
        "PICKED UP",
        "EXCEPTION",
        "DELAY",
        "DELIVERY ATTEMPT",
        "READY FOR CUSTOMER PICKUP",
        "TRACKING INFORMATION IS NOT YET AVAILABLE",
        "WE COULD NOT LOCATE THE SHIPMENT DETAILS",
        "INVALID TRACKING NUMBER",
    ))


def _ups_wait_for_batch(page, numbers, log):
    normal_deadline = time.time() + (UPS_TIMEOUT / 1000)
    human_deadline = None
    human_announced = False
    body = ""

    while True:
        try:
            body = page.locator("body").inner_text(timeout=5000)
        except Exception:
            body = ""

        upper = body.upper()
        numbers_present = any(number.upper() in upper for number in numbers)

        if numbers_present and _ups_result_phrases_present(body):
            if human_announced:
                log(
                    "[UPS] Human verification completed. "
                    "Tracking results detected; continuing automatically."
                )
            return body

        if _ups_challenge_present(body):
            if not human_announced:
                human_announced = True
                human_deadline = (
                    time.time()
                    + (UPS_HUMAN_VERIFICATION_TIMEOUT / 1000)
                )
                log(
                    "[UPS] Human verification is required in the "
                    "open Microsoft Edge window."
                )
                log(
                    "[UPS] Complete it manually and leave Edge open. "
                    "The program will continue automatically."
                )

            if time.time() >= human_deadline:
                raise RuntimeError(
                    "UPS human verification was not completed "
                    "within 10 minutes."
                )
        elif not human_announced and time.time() >= normal_deadline:
            raise RuntimeError(
                "UPS tracking results did not load within 60 seconds."
            )

        time.sleep(1)


def _ups_extract_results(page, body, numbers):
    results = {}

    for number in numbers:
        candidate = ""

        # First try to locate a DOM element containing this number and
        # inspect its nearest result card/container.
        try:
            locator = page.get_by_text(number, exact=False).first
            if locator.count():
                candidate = locator.evaluate(
                    """el => {
                        let node = el;
                        for (let i = 0; i < 10 && node; i++, node = node.parentElement) {
                            const t = (node.innerText || "").trim();
                            if (
                                t.length >= 20 &&
                                t.length <= 3000 &&
                                /Delivered|Out for Delivery|On the Way|In Transit|Label Created|We Have Your Package|Picked Up|Exception|Delay|Delivery Attempt|Ready for Customer Pickup|Tracking Information Is Not Yet Available|Invalid Tracking Number/i.test(t)
                            ) {
                                return t;
                            }
                        }
                        return (el.parentElement && el.parentElement.innerText) || el.innerText || "";
                    }"""
                )
        except Exception:
            candidate = ""

        # Fallback: inspect a text window around the tracking number.
        if not candidate:
            index = body.upper().find(number.upper())
            if index >= 0:
                candidate = body[
                    max(0, index - 250):
                    min(len(body), index + 2200)
                ]

        results[number] = _normalize_ups_status(candidate)

    return results


def _track_ups_batch(playwright, numbers, log):
    browser = playwright.chromium.launch(
        channel="msedge",
        headless=False,
        args=[
            "--disable-dev-shm-usage",
            "--start-maximized",
        ],
    )

    try:
        context = browser.new_context(
            locale="en-US",
            viewport=None,
        )
        page = context.new_page()
        page.set_default_timeout(30000)

        log(
            f"[UPS] Opening visible UPS tracking window for "
            f"{len(numbers)} number(s)."
        )
        log(
            "[UPS] Tracking numbers are being submitted automatically."
        )

        page.goto(
            _ups_tracking_url(numbers),
            wait_until="domcontentloaded",
            timeout=UPS_TIMEOUT,
        )

        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except PlaywrightTimeoutError:
            pass

        log(
            "[UPS] Waiting for UPS results or human verification..."
        )

        body = _ups_wait_for_batch(
            page,
            numbers,
            log,
        )

        results = _ups_extract_results(
            page,
            body,
            numbers,
        )

        usable = sum(
            1 for status in results.values()
            if status not in ("UNKNOWN", "")
        )
        log(
            f"[UPS] UPS page returned {usable} usable status(es)."
        )

        return results
    finally:
        browser.close()


def _update_ups(workbook, today, log):
    changed = False
    updated = unchanged = failed = 0
    references = []

    log("[UPS] Starting visible assisted UPS website update.")
    log(
        "[UPS] Processing current-month shipments "
        "in batches of up to 25."
    )

    for sheet_name in UPS_SHEETS:
        ws = _sheet(workbook, sheet_name)
        if ws is None:
            log(f"[UPS] {sheet_name}: sheet not found; skipped.")
            continue

        count = 0

        for row in range(1, ws.max_row + 1):
            carrier = str(ws.cell(row, COL_CARRIER).value or "")
            number = _tracking(ws.cell(row, COL_TRACKING).value)
            old = str(
                ws.cell(row, COL_REMARK).value or ""
            ).strip().upper()
            shipped = _as_date(ws.cell(row, COL_DATE).value)

            if not UPS_RE.search(carrier):
                continue

            # M/E is handled through the FedEx API.
            if re.fullmatch(r"\s*M\s*/\s*E\s*", carrier, re.I):
                continue

            if not number or not _current_month(shipped, today):
                continue

            if old == "DELIVERED":
                _set_status(ws, row, "DELIVERED")
                changed = True
                continue

            references.append((ws, sheet_name, row, number, old))
            count += 1

        log(f"[UPS] {sheet_name}: {count} eligible shipment(s).")

    if not references:
        return {
            "changed": changed,
            "updated": updated,
            "unchanged": unchanged,
            "failed": failed,
            "assisted": 0,
        }

    if sync_playwright is None:
        log("[UPS] Playwright is not available in this build.")
        return {
            "changed": changed,
            "updated": updated,
            "unchanged": unchanged,
            "failed": len(references),
            "assisted": 0,
        }

    unique_numbers = list(dict.fromkeys(item[3] for item in references))
    result_map = {}
    batch_count = 0

    with sync_playwright() as playwright:
        for batch_number, batch in enumerate(
            _batches(unique_numbers, UPS_BATCH_SIZE),
            start=1,
        ):
            batch_count += 1
            try:
                log(
                    f"[UPS] Batch {batch_number}: "
                    f"{len(batch)} tracking number(s)."
                )
                result_map.update(
                    _track_ups_batch(
                        playwright,
                        batch,
                        log,
                    )
                )
            except Exception as exc:
                log(
                    f"[UPS] Batch {batch_number} failed: "
                    f"{type(exc).__name__}: {exc}"
                )

    for ws, sheet_name, row, number, old in references:
        status = result_map.get(number)

        if not status or status == "UNKNOWN":
            failed += 1
            log(
                f"[UPS] {sheet_name} row {row}: "
                f"no usable status for {number}."
            )
            continue

        if status == old:
            unchanged += 1
            continue

        _set_status(ws, row, status)
        changed = True
        updated += 1

        log(
            f"[UPS] {sheet_name} row {row}: "
            f"{number} -> {status}"
        )

    return {
        "changed": changed,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
        "assisted": batch_count,
    }


def update_fedex_statuses(
    excel_path,
    *,
    environment="production",
    today=None,
    log=print,
):
    """
    Existing app.py can keep calling this function.
    It now updates both FedEx and USPS.
    """
    today = today or date.today()

    if not excel_path or not os.path.isfile(excel_path):
        message = f"[TRACKING] Workbook not found: {excel_path}"
        log(message)
        return {
            "updated": 0,
            "unchanged": 0,
            "failed": 0,
            "message": message,
        }

    log("[TRACKING] Starting combined FedEx + USPS + UPS + Malca-Amit update.")
    log(f"[TRACKING] Current month: {today.strftime('%B %Y')}")
    log("[TRACKING] Writing statuses to Remark column F.")

    workbook = _open_workbook(excel_path)
    try:
        fedex = _update_fedex(
            workbook,
            environment,
            today,
            log,
        )
        usps = _update_usps(
            workbook,
            today,
            log,
        )
        ups = _update_ups(
            workbook,
            today,
            log,
        )
        malca = _update_malca(
            workbook,
            today,
            log,
        )

        if (
            fedex["changed"]
            or usps["changed"]
            or ups["changed"]
            or malca["changed"]
        ):
            workbook.save(excel_path)
            log("[TRACKING] Workbook saved successfully.")
        else:
            log("[TRACKING] No workbook changes were required.")

        updated = fedex["updated"] + usps["updated"] + ups["updated"] + malca["updated"]
        unchanged = fedex["unchanged"] + usps["unchanged"] + ups["unchanged"] + malca["unchanged"]
        failed = fedex["failed"] + usps["failed"] + ups["failed"] + malca["failed"]

        message = (
            f"[TRACKING] Finished: updated={updated}, "
            f"unchanged={unchanged}, failed={failed}. "
            f"FedEx updated={fedex['updated']}; "
            f"USPS updated={usps['updated']}; "
            f"UPS updated={ups['updated']}; "
            f"UPS browser batches={ups['assisted']}; "
            f"Malca-Amit updated={malca['updated']}; "
            f"Malca-Amit needs-info={malca['needs_info']}."
        )
        log(message)

        return {
            "updated": updated,
            "unchanged": unchanged,
            "failed": failed,
            "fedex": fedex,
            "usps": usps,
            "ups": ups,
            "malca": malca,
            "message": message,
        }
    finally:
        workbook.close()


update_all_tracking_statuses = update_fedex_statuses


class FedExStatusScheduler(threading.Thread):
    """Existing name retained; scheduled runs now update FedEx, USPS, and Malca-Amit."""

    def __init__(
        self,
        get_excel_path,
        get_target,
        log=print,
        *,
        environment="production",
        poll_seconds=20,
    ):
        super().__init__(daemon=True)
        self.get_excel_path = get_excel_path
        self.get_target = get_target
        self.log = log
        self.environment = environment
        self.poll_seconds = poll_seconds
        self._stop_event = threading.Event()
        self._last_run_date = None

    def stop(self):
        self._stop_event.set()

    def run(self):
        self.log("FedEx + USPS + UPS + Malca-Amit status scheduler started.")

        while not self._stop_event.is_set():
            now = datetime.now()
            target = str(self.get_target() or "").strip()

            if (
                target
                and now.strftime("%H:%M") == target
                and self._last_run_date != now.date()
            ):
                self._last_run_date = now.date()
                path = str(self.get_excel_path() or "").strip()

                if path and os.path.isfile(path):
                    self.log(
                        "[TRACKING] Running scheduled "
                        "FedEx + USPS update..."
                    )
                    update_fedex_statuses(
                        path,
                        environment=self.environment,
                        today=now.date(),
                        log=self.log,
                    )
                else:
                    self.log(
                        "[TRACKING] Scheduled update skipped: "
                        "Excel path is missing."
                    )

            self._stop_event.wait(self.poll_seconds)

        self.log("FedEx + USPS + UPS + Malca-Amit status scheduler stopped.")
