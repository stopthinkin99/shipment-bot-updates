"""
daily_digest.py
---------------
Scheduled daily shipment-summary emails using Microsoft Graph delegated
authentication. Outlook desktop is not required.

This module:
  1. opens the tracking workbook,
  2. reads today's rows from UNI, EMBY, FENIX, and SOL,
  3. builds one HTML email per sheet, and
  4. sends each message to the email addresses configured below.

Requires:
    email_sender.py
    msal
    requests
    openpyxl
"""

import html
import json
import os
import shutil
import tempfile
import threading
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl

from email_sender import send_email


# ------------------------------------------------------------------ #
#  RECIPIENT CONFIGURATION
# ------------------------------------------------------------------ #
# Replace the example addresses with the real members of each group.
# Every address listed under a sheet receives that sheet's daily digest.
SHEET_TO_RECIPIENTS = {
    "UNI": [
        "Tenzin@unidesignusa.com",
        "tenzin.sundue@unidesignusa.com",
        "tsering@unidesignusa.com",
        "jimmy.huang@unidesignusa.com",
        "alex@unidesignusa.com",
        "hemi@unidesignusa.com",
    ],
    "EMBY": [
        "sujit.das@embyintl.com",
        "julia.napalkova@embyintl.com",
        "nirav.mehta@embyintl.com",
        "daniel.marin@embyintl.com",
        "hemant.jhaveri@embyintl.com",
    ],
    "FENIX": [
        "moe.khaing@fenixdiamonds.com",
        "lin.aung@fenixdiamonds.com",
        "gayatri.jariwala@fenixdiamonds.com",
        "shwe@fenixdiamonds.com",
        "sales@fenixdiamonds.com",
        "cesar.villar@fenixdiamonds.com",
        "hemant.jhaveri@fenixdiamonds.com",
    ],
    "SOL": [
        "shipping@unidesignusa.com",
    ],
}

# If True, sheets with no shipments today are skipped.
# If False, the configured recipients receive a "no shipments today" email.
SKIP_EMPTY_SHEETS = True


# ------------------------------------------------------------------ #
#  WORKBOOK COLUMN CONFIGURATION
# ------------------------------------------------------------------ #
# A DATE | B SHIP TO | C INVOICE/MEMO | D CARRIER | E TRACKING NO
IDX_DATE = 0
IDX_COMPANY = 1
IDX_INVOICE = 2
IDX_CARRIER = 3
IDX_TRACKING = 4


# ------------------------------------------------------------------ #
#  TIME PERSISTENCE
# ------------------------------------------------------------------ #
def _settings_path() -> Path:
    """
    Store the digest time in a user-writable location.

    Program Files is often read-only, so LOCALAPPDATA is preferred.
    """
    base = os.environ.get("LOCALAPPDATA")
    if base:
        folder = Path(base) / "UniCreation" / "ShipmentBot"
    else:
        folder = Path.home() / ".un_creation_shipment_bot"

    folder.mkdir(parents=True, exist_ok=True)
    return folder / "digest_settings.json"


_SETTINGS_FILE = _settings_path()


def load_digest_time(default="17:30"):
    """Return the saved HH:MM time, or `default` if none is available."""
    try:
        with _SETTINGS_FILE.open("r", encoding="utf-8") as file:
            value = json.load(file).get("digest_time", default)
            return str(value).strip() or default
    except Exception:
        return default


def save_digest_time(value):
    """Persist the HH:MM digest time."""
    try:
        with _SETTINGS_FILE.open("w", encoding="utf-8") as file:
            json.dump({"digest_time": str(value).strip()}, file, indent=2)
    except Exception:
        # A settings-write failure must never crash the GUI.
        pass


# ------------------------------------------------------------------ #
#  WORKBOOK READING
# ------------------------------------------------------------------ #
def _open_workbook(excel_path, retries=4, wait=1.5):
    """
    Open the workbook read-only.

    If Excel temporarily locks it, retry. If it remains locked, copy it
    to the temporary directory and read the copy.
    """
    for _ in range(retries):
        try:
            workbook = openpyxl.load_workbook(
                excel_path,
                read_only=True,
                data_only=True,
            )
            return workbook, None
        except (PermissionError, OSError):
            time.sleep(wait)

    temporary_path = os.path.join(
        tempfile.gettempdir(),
        f"digest_{os.path.basename(excel_path)}",
    )
    shutil.copy2(excel_path, temporary_path)

    workbook = openpyxl.load_workbook(
        temporary_path,
        read_only=True,
        data_only=True,
    )
    return workbook, temporary_path


def _as_date(value):
    """Convert an Excel or text value to datetime.date."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    formats = (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%d%b%y",
        "%d%b%Y",
    )

    for date_format in formats:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue

    return None


def _get(row, index):
    if index < len(row) and row[index] is not None:
        return str(row[index]).strip()
    return ""


def read_todays_shipments(worksheet, today):
    """Return all real shipment rows from `worksheet` dated `today`."""
    shipments = []

    for row in worksheet.iter_rows(values_only=True):
        if row is None or not any(cell is not None for cell in row):
            continue

        row_date = row[IDX_DATE] if len(row) > IDX_DATE else None
        if _as_date(row_date) != today:
            continue

        company = _get(row, IDX_COMPANY)
        tracking = _get(row, IDX_TRACKING)

        # Ignore divider or spacer rows.
        if not company and not tracking:
            continue

        shipments.append({
            "company": company,
            "invoice": _get(row, IDX_INVOICE),
            "tracking": tracking,
            "carrier": _get(row, IDX_CARRIER),
        })

    return shipments


def collect_by_sheet(excel_path, today=None):
    """
    Return:
        {
            "UNI": [...],
            "EMBY": [...],
            ...
        }

    Only configured sheets are considered.
    """
    today = today or date.today()
    workbook, temporary_path = _open_workbook(excel_path)
    result = {}

    try:
        actual_titles = {
            title.upper(): title
            for title in workbook.sheetnames
        }

        for sheet_key in SHEET_TO_RECIPIENTS:
            actual_title = actual_titles.get(sheet_key.upper())
            if not actual_title:
                continue

            shipments = read_todays_shipments(
                workbook[actual_title],
                today,
            )

            if shipments or not SKIP_EMPTY_SHEETS:
                result[sheet_key] = shipments
    finally:
        workbook.close()

        if temporary_path and os.path.exists(temporary_path):
            try:
                os.remove(temporary_path)
            except OSError:
                pass

    return result


# ------------------------------------------------------------------ #
#  EMAIL BUILDING AND SENDING
# ------------------------------------------------------------------ #
def _safe(value):
    return html.escape(str(value or ""))


def _build_html(sheet_key, shipments, today):
    day = today.strftime("%m/%d/%Y")

    if not shipments:
        return (
            "<p style='font-family:Segoe UI,Arial,sans-serif'>"
            f"No {_safe(sheet_key)} shipments went out on {_safe(day)}."
            "</p>"
        )

    header = (
        "<tr style='background:#5B9BD5;color:#ffffff'>"
        "<th align='left' style='padding:6px'>Company</th>"
        "<th align='left' style='padding:6px'>Invoice</th>"
        "<th align='left' style='padding:6px'>Tracking #</th>"
        "<th align='left' style='padding:6px'>Carrier</th>"
        "</tr>"
    )

    body_rows = "".join(
        "<tr>"
        f"<td style='padding:6px'>{_safe(shipment['company'])}</td>"
        f"<td style='padding:6px'>{_safe(shipment['invoice'])}</td>"
        f"<td style='padding:6px'>{_safe(shipment['tracking'])}</td>"
        f"<td style='padding:6px'>{_safe(shipment['carrier'])}</td>"
        "</tr>"
        for shipment in shipments
    )

    return (
        "<div style='font-family:Segoe UI,Arial,sans-serif;font-size:13px'>"
        f"<p>{len(shipments)} {_safe(sheet_key)} shipment(s) "
        f"went out on {_safe(day)}:</p>"
        "<table border='1' cellpadding='0' cellspacing='0' "
        "style='border-collapse:collapse'>"
        f"{header}{body_rows}"
        "</table>"
        "<p>This is an automated message from the "
        "Uni Creation Shipment Bot.</p>"
        "</div>"
    )


def send_group_digest(
    sheet_key,
    shipments,
    today,
    outlook=None,  # retained for compatibility with older callers
    log=print,
):
    """
    Send one digest through Microsoft Graph.

    `outlook` is intentionally ignored. It remains in the signature so
    older code that passes it does not break.
    """
    del outlook

    recipients = [
        address.strip()
        for address in SHEET_TO_RECIPIENTS.get(sheet_key, [])
        if address and address.strip()
    ]

    if not recipients:
        return False, f"{sheet_key}: no recipient email addresses configured"

    subject = (
        f"Shipments Sent Today - {sheet_key} - "
        f"{today.strftime('%m/%d/%Y')}"
    )

    send_email(
        recipients=recipients,
        subject=subject,
        body=_build_html(sheet_key, shipments, today),
        html=True,
        log_fn=log,
    )

    return (
        True,
        f"{sheet_key}: sent to {len(recipients)} recipient(s) "
        f"({len(shipments)} shipment row(s))",
    )


def run_daily_digest(excel_path, today=None, log=print):
    """
    Read today's shipments and send one message per configured sheet.

    Returns a list of status messages.
    """
    today = today or date.today()

    if not excel_path:
        message = "Digest failed: Excel path is empty."
        log(message)
        return [message]

    if not os.path.isfile(excel_path):
        message = f"Digest failed: Excel file not found: {excel_path}"
        log(message)
        return [message]

    try:
        by_sheet = collect_by_sheet(excel_path, today)
    except Exception as exc:
        message = (
            "Digest read failed: "
            f"{type(exc).__name__}: {exc}"
        )
        log(message)
        return [message]

    if not by_sheet:
        message = (
            "No shipments found for "
            f"{today.strftime('%m/%d/%Y')} on any configured sheet."
        )
        log(message)
        return [message]

    results = []

    for sheet_key, shipments in by_sheet.items():
        try:
            _, information = send_group_digest(
                sheet_key,
                shipments,
                today,
                log=log,
            )
        except Exception as exc:
            information = (
                f"{sheet_key}: FAILED — "
                f"{type(exc).__name__}: {exc}"
            )

        log(information)
        results.append(information)

    return results


# ------------------------------------------------------------------ #
#  SCHEDULER
# ------------------------------------------------------------------ #
class DigestScheduler(threading.Thread):
    """
    Background thread that runs the digest once per day at HH:MM.
    """

    def __init__(
        self,
        get_excel_path,
        get_target,
        log=print,
        poll_seconds=20,
    ):
        super().__init__(daemon=True)
        self.get_excel_path = get_excel_path
        self.get_target = get_target
        self.log = log
        self.poll_seconds = poll_seconds
        self._stop_event = threading.Event()
        self._last_fired = None

    def stop(self):
        self._stop_event.set()

    def run(self):
        self.log("Daily digest scheduler started.")

        while not self._stop_event.is_set():
            target = (self.get_target() or "").strip()
            now = datetime.now()

            should_run = (
                target
                and now.strftime("%H:%M") == target
                and self._last_fired != now.date()
            )

            if should_run:
                self._last_fired = now.date()
                excel_path = self.get_excel_path()

                if not excel_path or not os.path.exists(excel_path):
                    self.log(
                        "Digest skipped: Excel path is not set or missing."
                    )
                else:
                    self.log(
                        "Running daily digest for "
                        f"{now.strftime('%m/%d/%Y')} ..."
                    )
                    run_daily_digest(
                        excel_path,
                        today=now.date(),
                        log=self.log,
                    )

            self._stop_event.wait(self.poll_seconds)

        self.log("Daily digest scheduler stopped.")